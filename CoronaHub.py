import tkinter as tk
from tkinter import *
import tkinter.ttk as ttk
from PIL import ImageTk, Image
import openpyxl
import tweepy
from newsapi import NewsApiClient
import pandas as pd
import numpy as np 


#Setup
root =Tk()
root.title("Corona Info")
root.iconbitmap("medical_icon.ico")
root.geometry("600x400")

#Variables
rating = IntVar()
rating_string = StringVar()
USRS_LIST = ["BNODesk", "Breaking911", "DeItaOne", "MisterAntiBully"]
twitter = StringVar()
twitter.set("BNODesk")
news = StringVar()
state = StringVar()
global state_names
state_names = ["Alaska", "Alabama", "Arkansas", "American Samoa", "Arizona", "California", "Colorado", "Connecticut", "District ", "of Columbia", "Delaware", "Florida", "Georgia", "Guam", "Hawaii", "Iowa", "Idaho", "Illinois", "Indiana", "Kansas", "Kentucky", "Louisiana", "Massachusetts", "Maryland", "Maine", "Michigan", "Minnesota", "Missouri", "Mississippi", "Montana", "North Carolina", "North Dakota", "Nebraska", "New Hampshire", "New Jersey", "New Mexico", "Nevada", "New York", "Ohio", "Oklahoma", "Oregon", "Pennsylvania", "Puerto Rico", "Rhode Island", "South Carolina", "South Dakota", "Tennessee", "Texas", "Utah", "Virginia", "Virgin Islands", "Vermont", "Washington", "Wisconsin", "West Virginia", "Wyoming"]

#scrollable frame class if needed
class ScrollableFrame(ttk.Frame):
    def __init__(self, container, *args, **kwargs):
        super().__init__(container, *args, **kwargs)
        canvas = tk.Canvas(self)
        scrollbar = ttk.Scrollbar(self, orient="vertical", command=canvas.yview)
        self.scrollable_frame = ttk.Frame(canvas)

        self.scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(
                scrollregion=canvas.bbox("all")
            )
        )

        canvas.create_window((0, 0), window=self.scrollable_frame, anchor="nw")

        canvas.configure(yscrollcommand=scrollbar.set)

        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

#Get Tweets Function
def pull_tweets(twitter_name):
    consumer_key = 'VgVDRfXxvyAGNSGQm260yTRx6'
    consumer_secret = 'GuW8Kw5P2WkgRRp15mDh5kZ8JJkebbTJ9UgxDClQL23HGdKDtM'
    access_key = '756186848-bYXgiENpn4qvWFJivxolimfB4VedWJYyaLCeKETJ'
    access_secret = '3tiS5wSeFO2unbXE7RFqGoC1NHiARDaNMr7mDJUyEzBwC'
    auth = tweepy.OAuthHandler(consumer_key, consumer_secret)
    auth.set_access_token(access_key, access_secret)
    api = tweepy.API(auth, wait_on_rate_limit=True)
    try:
        api.verify_credentials()
    except: return

    corona_tweets = api.user_timeline(screen_name=twitter_name, count=10)
    text = " "
    
    for status in corona_tweets:
        temp_string = str(status.text)
        text = text + "\n" + "\n" + temp_string

    return text


def by_state(state):
    corona_df = pd.read_csv("corona_data.csv")
    corona_df = corona_df.drop("latitude",axis=1,)
    corona_df = corona_df.drop("longitude",axis=1,)
    corona_df = corona_df.drop("recovered",axis=1,)
    corona_df = corona_df[corona_df.countryregion == "US" ]
    corona_df = corona_df.sort_values(by=['provincestate'])
    corona_df = corona_df[corona_df.provincestate == state]
    text = corona_df.to_string()
    if corona_df.empty:
        return "NO DATA AVAILABLE BECAUSE OPENBLENDER COSTS MONEY"
    else: return text

#DATA HUB
def goto_data_hub():
    data_hub=Toplevel()
    data_hub_title = Label(data_hub,text="Data Hub",font=("Helvetica", 72))
    state_picker = OptionMenu(data_hub,state,*state_names )
    data_label = Label(data_hub, text = "State Data")


    data_hub_title.grid(row=0, column=0)
    state_picker.grid(row=1,column=0)
    data_label.grid(row=2, column=0)

    def change_dropdown(*args):
        temp = by_state(state.get())
        data_label = Label(data_hub, text = temp)
        data_label.grid(row=2, column=0)
    state.trace('w',change_dropdown)
    return

#NEWS HUB
def goto_news_hub():
    news_hub=Toplevel()
    news_hub.geometry("1400x1080")

    def get_the_news ():
        api = NewsApiClient(api_key='ec63b55e37e546118ac2a7dffd55239e')
        corona_news = api.get_top_headlines(q='coronavirus',category="health", country="us", page_size=15)
        corona_news = corona_news['articles']
        results = []
        final = []
        text = '\n \n'
        for i in corona_news:
            results.append(i["title"])
        for a in range(len(results)): 
            final.append(results[a])
        news = text.join(final)
        news_label = Label(news_frame, text = news, width=90, height=40)
        news_label.grid(row=2, column=0)
        return
    
    news_hub_title = Label(news_hub,text="News Hub",font=("Helvetica", 72))
    news_frame = LabelFrame(news_hub, padx = 10)
    twitter_frame = LabelFrame(news_hub, padx = 10)
    news_frame_title = Label(news_frame, text = "Latest Headlines" ,font=("Helvetica", 18))
    twitter_frame_title = Label(twitter_frame, text = "Tweets",font=("Helvetica", 18))
    twitter_user_choice = OptionMenu(twitter_frame, twitter, *USRS_LIST)
    tweets = Label(twitter_frame, text = "Tweets go here!", width=95, height=40)
    news_button = Button(news_frame,text = 'Load Latest Headlines',font=("Helvetica", 12), command = get_the_news())
    news_label = Label(news_frame, text ="News Here")


    news_hub_title.grid(row=0,column=0,columnspan=2)
    news_frame.grid(row=1, column=0)
    news_frame_title.grid(row=0, column=0)
    news_button.grid(row=1, column=0)
    news_label.grid(row=2, column=0)
    twitter_frame.grid(row=1, column=1)
    twitter_frame_title.grid(row=0, column=0)
    twitter_user_choice.grid(row=1, column=0)
    tweets.grid(row=2, column=0)
    
    def change_dropdown(*args):
        t_temp = pull_tweets(twitter.get())
        tweets['text'] = t_temp
        tweets.grid(row=2, column=0)

    twitter.trace('w',change_dropdown)
    return

#This function takes the values from the rating radio buttons 
# and suggestion text box and places them in an excel file. 0 is good, 1 is bad due to unresolved radio button issues. 
# It also resets the display and removes button usability
def clicked_rating_button(var,string):
    rating_workbook = openpyxl.load_workbook('AppRatings.xlsx')
    ratings = rating_workbook["Sheet1"]
    count = ratings.max_row + 1
    ratings.cell(row=count, column=1).value = rating.get()
    ratings.cell(row=count, column=2).value = rating_string.get()
    rating_workbook.save('AppRatings.xlsx')
    
    #RESET DISPLAY
    b_rating = Button(rating_system_frame,state=DISABLED, text="Send",padx=5,pady=5,command=lambda: clicked_rating_button(rating.get(),rating_string.get()))
    rating_text.delete(0,END)
    rating_text.insert(0, "Thank you for your feedback!")
    b_rating.grid(row=3, columnspan=2)
    return

#Main Page

coronainfo_label = Label(root,text="Corona Info",font=("Helvetica", 72), padx=50)
main_page_button_frame = LabelFrame(root,padx=89,pady=10)
main_page_label = Label(main_page_button_frame, text="Click Here To Go To Modules",font=("Helvetica", 11))
datahub_button = Button(main_page_button_frame,text="Data Hub",padx=5,pady=5,command=goto_data_hub)
newshub_button = Button(main_page_button_frame,text="News Hub",padx=5,pady=5,command=goto_news_hub )

#App rating system

rating_system_frame = LabelFrame(root,padx=60,pady=10)
rating_label = Label(rating_system_frame,text="How Do You Like This App?",font=("Helvetica", 11))
rb_good = Radiobutton(rating_system_frame, text="Good", variable=rating, value=0, state=NORMAL)
rb_bad = Radiobutton(rating_system_frame, text="Bad", variable=rating, value=1, state=NORMAL)
rating_text = Entry(rating_system_frame,width=40,textvariable=rating_string,relief=SUNKEN)
b_rating = Button(rating_system_frame, text="Send",padx=5,pady=5,command=lambda: clicked_rating_button(rating.get(),rating_string.get()))



#Main Page Display

coronainfo_label.grid(row=0,column=0,columnspan=2)
main_page_button_frame.grid(row=1,columnspan=2)
main_page_label.grid(row=0,column=0, columnspan=2)
datahub_button.grid(row=1,column=0)
newshub_button.grid(row=1,column=1)

rating_system_frame.grid(row=2,columnspan=2)
rating_label.grid(row=0,columnspan=2)
rb_good.grid(row=1, column=0)
rb_bad.grid(row=1, column=1)
rating_text.grid(row=2,column=0,columnspan=2,padx=5,pady=10,ipady=3)
rating_text.insert(0, "Suggestions?")
b_rating.grid(row=3, columnspan=2)




root.mainloop()